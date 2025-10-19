# app.py
# =======================================================================================
# 목적: 하나의 앱에서 ① 관내출장여비, ② 초과근무내역, ③ 자료 수합(머릿글 유지 수합)을 처리
#
# [탭 안내 · 상세 로직 주석]
# ───────────────────────────────────────────────────────────────────────────────────────
# ■ ① 관내출장여비
#   1) 업로드용 백데이터 준비
#      - (인사랑) 원본(.xlsx)과 (서식) 출장자 백데이터(.xlsx) 안내/다운로드 제공
#   2) 파일 업로드
#      - 원본 및 (서식) 업로드
#   3) 데이터 가공·요약
#      - 백데이터 시트 생성(병합 해제, 여분 삭제, 빈 이름 행 삭제)
#      - DataFrame 변환 → 규칙 적용(4시간 구분, 1시간 미만, 지급단가 결정) → "가공" 시트 작성
#      - UI에서 연/월/부서 선택, 특정 출장자/단가별 날짜 ‘제외/포함’ 규칙 누적 → 요약표 생성
#   4) 지급 조서 생성·다운로드
#      - 요약표 + (서식) 백데이터를 결합해 혼합 DF 작성(20,000/10,000 블록 보장)
#      - ‘혼합’ 시트로 출력
#      - 서식 후처리(머릿글 병합, 금액서식, 합계열 삽입, 총합계, 푸터, 열너비 자동 등)
#
# ■ ② 초과근무내역
#   1) 업로드용 백데이터 준비
#      - (서식) 초과근무자 백데이터 안내/다운로드 제공
#   2) 파일 업로드
#      - (서식) 업로드
#   3) 데이터 가공·요약
#      - 기준 연/분기 입력
#      - 분기 월(3개월)별 수당시간을 월57h/분기90h 상한 규칙으로 보정
#      - 강제조정 비고(월57h/분기90h 사유) 포함, 누계 강조, 57h 표시
#   4) 엑셀 저장
#      - 화면 표기 그대로 엑셀로 저장(머릿글, 테두리, 열너비, 고정창 등 적용)
#
# ■ ③ 자료 수합(머릿글 유지 수합 도구)
#   1) 여러 엑셀 업로드(xls/xlsx)
#   2) 머릿글 범위 설정(첫 행~마지막 행, 1-based)  ※ 기본: 1~1
#   3) 데이터 범위 설정(시작 행~마지막 행 또는 끝까지, 1-based)
#   4) 수합·정규화
#      - 머릿글 블록 ffill로 병합 흔적 평면화
#      - 우측 연속 공란 제거로 “실사용 열 너비”만 사용
#      - 컬럼명은 위→아래 텍스트를 '_'로 결합, 중복명 _1, _2 부여
#      - 본문 뒤에 ‘출처’ 열 추가
#   5) 대표 파일의 머릿글 병합 모양을 상대좌표로 캡처
#   6) 결과 엑셀 생성
#      - 병합 머릿글 복원 + ‘출처’ 머릿글도 동일 높이로 병합
#      - 본문 기입, ‘출처’ 열 연한 파랑, 열 너비 추정
#   7) 다운로드
#      - 파일명: “수합 완료본_yymmdd_hhmm.xlsx”(KST)
# =======================================================================================

from __future__ import annotations

import os
import re
from io import BytesIO
from typing import Optional
from datetime import datetime, timedelta, timezone

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# ----------------------------------
# 앱 상수
# ----------------------------------
APP_TITLE = "관내출장여비 · 초과근무내역 · 자료 수합"
MANUAL_FILE = "인사랑 관내출장 내역 추출.pdf"
FORM_TEMPLATE_FILE = "(서식) 출장자 백데이터.xlsx"
FORM_TEMPLATE_FILE_OVT = "(서식) 초과근무자 백데이터.xlsx"

TARGET_HEADERS = ["순번", "출장자", "도착일자", "총출장시간", "차량",
                  "4시간구분", "1시간미만", "지급단가", "여비금액"]
REQUIRED_SRC = ["순번", "출장자", "도착일자", "총출장시간", "차량"]

FILL_HEADER = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")  # 연한 파랑
THIN_SIDE = Side(style="thin", color="000000")
BORDER_THIN = Border(top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
PINK = PatternFill(fill_type="solid", start_color="FFC0CB", end_color="FFC0CB")          # 연한 분홍

OVT_MONTH_CAP = 57.0
OVT_QTR_CAP = 90.0

# ----------------------------------
# 시간대(KST)
# ----------------------------------
try:
    from zoneinfo import ZoneInfo
    KST = ZoneInfo("Asia/Seoul")
except ImportError:
    from pytz import timezone as _tz
    KST = _tz("Asia/Seoul")

def kst_timestamp() -> str:
    return datetime.now(KST).strftime("%y%m%d_%H%M")

def _kst_now() -> datetime:
    return datetime.now(timezone(timedelta(hours=9)))

# ----------------------------------
# 규칙/판정 보조 상수·함수(출장)
# ----------------------------------
_HOURS_GE4 = set(map(str, range(4, 24)))
_HOURS_LT4 = {"1", "2", "3"}

def _extract_hour_token(s: str) -> str | None:
    m = re.search(r"(\d+)\s*시간", s)
    return m.group(1) if m else None

def rule_4h_bucket(s: str) -> str:
    s = "" if pd.isna(s) else str(s)
    s = s.replace(" ", "")
    has_day, has_hour, has_min = ("일" in s), ("시간" in s), ("분" in s)
    if has_day:
        return "4시간이상"
    if has_hour and has_min:
        h = _extract_hour_token(s)
        if h in _HOURS_GE4:
            return "4시간이상"
        if h in _HOURS_LT4:
            return "4시간미만"
        return "4시간미만"
    if has_hour and not has_min:
        h = _extract_hour_token(s)
        if h in _HOURS_GE4:
            return "4시간이상"
        if h in _HOURS_LT4:
            return "4시간미만"
        return ""
    if (not has_hour) and (not has_day) and has_min:
        return "4시간미만"
    return ""

def rule_under1h(s: str) -> str:
    s = "" if pd.isna(s) else str(s)
    s = s.replace(" ", "")
    return "1시간미만" if ("시간" not in s and "일" not in s) and ("분" in s) else ""

def rule_pay(x_val: str, car_val: str) -> int:
    x = (x_val or "").strip()
    k = (car_val or "").strip()
    if x == "4시간이상" and k == "미사용":
        return 20000
    if x == "4시간이상" and k == "사용":
        return 10000
    if x == "4시간미만" and k == "미사용":
        return 10000
    if x == "4시간미만" and k == "사용":
        return 0
    return 0

# ----------------------------------
# DataFrame/엑셀 유틸
# ----------------------------------
def to_datetime_flex(v):
    if pd.isna(v):
        return pd.NaT
    if isinstance(v, (datetime, pd.Timestamp)):
        return pd.to_datetime(v)
    try:
        if isinstance(v, (int, float)) or (isinstance(v, str) and v.replace(".", "", 1).isdigit()):
            num = float(v)
            base = datetime(1899, 12, 30)
            return pd.to_datetime(base + timedelta(days=num))
    except Exception:
        pass
    try:
        return pd.to_datetime(str(v), errors="coerce")
    except Exception:
        return pd.NaT

def ws_to_dataframe(ws: Worksheet) -> pd.DataFrame:
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    header = [("" if v is None else str(v).strip()) for v in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)

def prepare_backend_sheet_xlsx(file_like):
    wb = load_workbook(file_like)
    ws = wb.active
    ws.title = "백데이터"
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.delete_cols(1, 1)
    ws.delete_rows(1, 3)
    for r in range(ws.max_row, 2, -1):
        v = ws.cell(row=r, column=3).value
        if v is None or str(v).strip() == "":
            ws.delete_rows(r, 1)
    return wb

def save_wb_to_bytes(wb) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def read_template_dataframe(file_like) -> pd.DataFrame:
    wb = load_workbook(file_like, data_only=True)
    ws = wb.active
    rows = list(ws.values)
    wb.close()
    if not rows:
        return pd.DataFrame()
    header = [("" if v is None else str(v).strip()) for v in rows[0]]
    df = pd.DataFrame(rows[1:], columns=header).dropna(how="all")
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].apply(lambda x: "" if x is None else str(x).strip())
    return df

# ----------------------------------
# 가공/요약 생성(출장)
# ----------------------------------
def create_gagong_and_summary(wb):
    dfb = ws_to_dataframe(wb["백데이터"])
    missing = [c for c in REQUIRED_SRC if c not in dfb.columns]
    if missing:
        raise RuntimeError(f"백데이터 필수 열 누락: {', '.join(missing)}")

    seq = dfb["순번"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    name = dfb["출장자"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    arrv_dt = dfb["도착일자"].apply(to_datetime_flex)
    time_str = dfb["총출장시간"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    car = dfb["차량"].apply(lambda x: "" if pd.isna(x) else str(x).strip())

    proc = pd.DataFrame({
        "순번": seq,
        "출장자": name,
        "도착일자": arrv_dt.dt.strftime("%Y-%m-%d"),
        "총출장시간": time_str,
        "차량": car,
    })
    proc["4시간구분"] = proc["총출장시간"].apply(rule_4h_bucket)
    proc["1시간미만"] = proc["총출장시간"].apply(rule_under1h)
    proc["지급단가"] = proc.apply(lambda r: rule_pay(r["4시간구분"], r["차량"]), axis=1)
    proc["여비금액"] = proc["지급단가"]
    proc = proc[TARGET_HEADERS]

    if "가공" in wb.sheetnames:
        del wb["가공"]
    ws_p = wb.create_sheet("가공")
    ws_p.append(TARGET_HEADERS)
    for _, row in proc.iterrows():
        ws_p.append(list(row.values))

    if "요약" in wb.sheetnames:
        del wb["요약"]
    wb.create_sheet("요약").append(["출장자", "지급단가", "출장일수", "여비합계", "출장현황"])

    return wb, proc

# ----------------------------------
# 혼합 DF 생성(출장)
# ----------------------------------
def find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cols = {str(c).strip(): c for c in df.columns}
    for name in candidates:
        if name in cols:
            return cols[name]
    norm = {str(c).replace(" ", ""): c for c in df.columns}
    for name in candidates:
        key = name.replace(" ", "")
        if key in norm:
            return norm[key]
    return None

def parse_days(txt: str) -> list:
    if pd.isna(txt) or str(txt).strip() == "":
        return []
    tokens = [t.strip().replace("일", "") for t in str(txt).split(",")]
    out = []
    for t in tokens:
        if t == "":
            continue
        try:
            out.append(int(t))
        except Exception:
            out.append(str(t))
    nums = sorted([d for d in out if isinstance(d, int)])
    strs = [d for d in out if not isinstance(d, int)]
    return nums + strs

def _norm_serial(v):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return pd.NA
    n = pd.to_numeric(v, errors="coerce")
    return pd.NA if pd.isna(n) else int(float(n))

def build_mixed_df(summary_df: pd.DataFrame, tmpl_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df is None or summary_df.empty:
        raise RuntimeError("요약 표 데이터가 없습니다.")
    if tmpl_df is None or tmpl_df.empty:
        raise RuntimeError("(서식) 출장자 백데이터가 없습니다.")

    sdf = summary_df.copy()
    if "성명" not in sdf.columns and "출장자" in sdf.columns:
        sdf = sdf.rename(columns={"출장자": "성명"})
    for c in ["성명", "지급단가", "출장현황", "출장일수", "여비합계"]:
        if c not in sdf.columns:
            raise RuntimeError(f"요약 표에 '{c}' 열이 없습니다.")

    sdf["성명"] = sdf["성명"].astype(str).str.strip()
    sdf["지급단가"] = pd.to_numeric(sdf["지급단가"], errors="coerce").fillna(0).astype(int)
    sdf["출장일수"] = pd.to_numeric(sdf["출장일수"], errors="coerce").fillna(0).astype(int)
    sdf["여비합계"] = pd.to_numeric(sdf["여비합계"], errors="coerce").fillna(0).astype(int)
    sdf["__days_list__"] = sdf["출장현황"].apply(parse_days)

    by_key: dict[tuple[str, int], dict] = {}
    for _, r in sdf.iterrows():
        by_key[(r["성명"], int(r["지급단가"]))] = {
            "days": list(r["__days_list__"]),
            "cnt": int(r["출장일수"]),
            "sum": int(r["여비합계"]),
        }

    serial_col = find_col(tmpl_df, ["연번", "순번", "번호"])
    nm_col = find_col(tmpl_df, ["성명", "출장자"])
    rank_col = find_col(tmpl_df, ["직급", "직 급"])
    bank_col = find_col(tmpl_df, ["은행명", "은행"])
    acct_col = find_col(tmpl_df, ["계좌번호", "계좌"])
    if nm_col is None:
        raise RuntimeError("백데이터에서 성명/출장자 열을 찾지 못했습니다.")

    rows, max_days = [], 0
    TIERS = [20000, 10000]

    for _, row in tmpl_df.iterrows():
        nm = str(row.get(nm_col, "")).strip()
        if not nm:
            continue
        meta = {
            "연번": _norm_serial(row.get(serial_col, pd.NA)),
            "직급": str(row.get(rank_col, "") if rank_col else "").strip(),
            "성명": nm,
            "은행명": str(row.get(bank_col, "") if bank_col else "").strip(),
            "계좌번호": str(row.get(acct_col, "") if acct_col else "").strip(),
        }
        for pay in TIERS:
            rec = by_key.get((nm, pay), {"days": [], "cnt": 0, "sum": 0})
            days_list = list(rec["days"])
            max_days = max(max_days, len(days_list))
            rows.append({
                **meta,
                "__days__": days_list,
                "출장일수": int(rec["cnt"]) if rec["cnt"] else len(days_list),
                "지급단가": int(pay),
                "소계": int(rec["sum"]) if rec["sum"] else int(pay) * len(days_list),
            })

    date_cols = ["출장현황"] + [f"출장현황{i}" for i in range(2, max_days + 1)] if max_days > 0 else ["출장현황"]

    out_rows = []
    for r in rows:
        base = {k: r[k] for k in ["연번", "직급", "성명", "은행명", "계좌번호"]}
        for i in range(max_days):
            key = "출장현황" if i == 0 else f"출장현황{i+1}"
            base[key] = r["__days__"][i] if i < len(r["__days__"]) else ""
        base["출장일수"] = r["출장일수"]
        base["지급단가"] = r["지급단가"]
        base["소계"] = r["소계"]
        out_rows.append(base)

    cols = ["연번", "직급", "성명", "은행명", "계좌번호"] + date_cols + ["출장일수", "지급단가", "소계"]
    out_df = pd.DataFrame(out_rows, columns=cols)

    if "연번" in out_df.columns:
        out_df["연번"] = pd.to_numeric(out_df["연번"], errors="coerce").astype("Int64")

    return out_df

# ----------------------------------
# 엑셀 서식 유틸
# ----------------------------------
def set_alignment(ws: Worksheet, rows: range, cols: range, horizontal="center", vertical="center"):
    for rr in rows:
        for cc in cols:
            ws.cell(rr, cc).alignment = Alignment(horizontal=horizontal, vertical=vertical)

def set_row_border(ws: Worksheet, row: int, max_col: int, border: Border):
    for c in range(1, max_col + 1):
        ws.cell(row, c).border = border

def set_header_fill(ws: Worksheet, row: int, max_col: int, fill: PatternFill):
    for c in range(1, max_col + 1):
        ws.cell(row, c).fill = fill

def auto_col_width(ws: Worksheet):
    for c in range(1, ws.max_column + 1):
        max_len = 0
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(rr, c).value
            lv = len(str(v)) if v is not None else 0
            if lv > max_len:
                max_len = lv
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 60)

# ----------------------------------
# 혼합 DF → 엑셀 렌더링(출장)
# ----------------------------------
def export_mixed_to_excel(df: pd.DataFrame, year: int | None, month: int | None, dept: str | None) -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="혼합", index=False, startrow=4)
        ws = writer.book["혼합"]

        header_row = 5
        data_start = header_row + 1

        # 출장현황 헤더 병합
        first_status_col, last_status_col = None, None
        for c in range(1, ws.max_column + 1):
            h = ws.cell(header_row, c).value
            if isinstance(h, str) and h.startswith("출장현황"):
                if first_status_col is None:
                    first_status_col = c
                last_status_col = c
        if first_status_col and last_status_col and last_status_col > first_status_col:
            ws.merge_cells(start_row=header_row, start_column=first_status_col,
                           end_row=header_row, end_column=last_status_col)
            ws.cell(header_row, first_status_col).value = "출장현황"

        # 합계 열 삽입
        hdr_idx = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)}
        sub_col = hdr_idx.get("소계")
        if not sub_col:
            raise RuntimeError("'소계' 열을 찾을 수 없습니다.")
        total_col = sub_col + 1
        ws.insert_cols(total_col, amount=1)
        ws.cell(header_row, total_col).value = "합계"
        ws.cell(header_row, total_col).font = Font(bold=True)

        # 단위
        unit_row = header_row - 1
        ws.cell(unit_row, total_col).value = "(단위 : 원)"
        ws.cell(unit_row, total_col).alignment = Alignment(horizontal="right", vertical="center")

        # 제목
        title = f"{(dept or '').strip()} 관내 출장여비 지급내역({year or ''}년 {month or ''}월)"
        ws["A2"] = title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_col)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(size=20)

        # 인덱스 재구성
        hdr_idx = {ws.cell(header_row, c).value: c for c in range(1, ws.max_column + 1)}
        col_serial = hdr_idx.get("연번")
        col_rank = hdr_idx.get("직급")
        col_name = hdr_idx.get("성명")
        col_bank = hdr_idx.get("은행명")
        col_acct = hdr_idx.get("계좌번호")
        col_cnt = hdr_idx.get("출장일수")
        col_pay = hdr_idx.get("지급단가")
        col_sub = hdr_idx.get("소계")
        col_total = hdr_idx.get("합계")
        last_row = ws.max_row
        last_col = ws.max_column

        set_header_fill(ws, header_row, last_col, FILL_HEADER)

        # 동일 인적사항 블록 처리
        r = data_start
        while r <= last_row:
            key = (
                ws.cell(r, col_serial).value if col_serial else "",
                ws.cell(r, col_rank).value,
                ws.cell(r, col_name).value,
                ws.cell(r, col_bank).value,
                ws.cell(r, col_acct).value,
            )
            run_end = r
            while run_end + 1 <= last_row:
                k2 = (
                    ws.cell(run_end + 1, col_serial).value if col_serial else "",
                    ws.cell(run_end + 1, col_rank).value,
                    ws.cell(run_end + 1, col_name).value,
                    ws.cell(run_end + 1, col_bank).value,
                    ws.cell(run_end + 1, col_acct).value,
                )
                if k2 == key:
                    run_end += 1
                else:
                    break

            # 출장일수 = COUNTA(현황 구간)
            if first_status_col is not None and last_status_col is not None and col_cnt:
                sL = get_column_letter(first_status_col)
                eL = get_column_letter(last_status_col)
                for rr in range(r, run_end + 1):
                    cnt_cell = ws.cell(rr, col_cnt)
                    cnt_cell.value = f"=COUNTA({sL}{rr}:{eL}{rr})"
                    cnt_cell.number_format = "0"
                    cnt_cell.alignment = Alignment(horizontal="center", vertical="center")

            # 지급단가 서식
            for rr in range(r, run_end + 1):
                pay_cell = ws.cell(rr, col_pay)
                pay_cell.number_format = "#,##0"
                pay_cell.alignment = Alignment(horizontal="right", vertical="center")

            # 소계 = 출장일수 * 지급단가
            for rr in range(r, run_end + 1):
                sub_cell = ws.cell(rr, col_sub)
                cnt_cell = ws.cell(rr, col_cnt)
                pay_cell = ws.cell(rr, col_pay)
                sub_cell.value = f"={cnt_cell.coordinate}*{pay_cell.coordinate}"
                sub_cell.number_format = "#,##0"
                sub_cell.alignment = Alignment(horizontal="right", vertical="center")

            # 합계 = 블록 소계 합
            sub_coords = [ws.cell(rr, col_sub).coordinate for rr in range(r, run_end + 1)]
            total_formula = "=" + "+".join(sub_coords) if sub_coords else "=0"

            # 인적사항 병합
            to_merge = [x for x in [col_serial, col_rank, col_name, col_bank, col_acct, col_total] if x]
            if run_end > r:
                for c in to_merge:
                    ws.merge_cells(start_row=r, start_column=c, end_row=run_end, end_column=c)
                    ws.cell(r, c).alignment = Alignment(vertical="center", horizontal="center")

            # 합계 셀
            ws.cell(r, col_total).value = total_formula
            ws.cell(r, col_total).number_format = "#,##0"
            ws.cell(r, col_total).alignment = Alignment(horizontal="right", vertical="center")

            r = run_end + 1

        # 출장현황 가운데 정렬
        if first_status_col and last_status_col:
            for rr in range(data_start, last_row + 1):
                for cc in range(first_status_col, last_status_col + 1):
                    ws.cell(rr, cc).alignment = Alignment(horizontal="center", vertical="center")

        # 총합계
        last_data_row = ws.max_row
        totals_row = last_data_row + 1
        ws.cell(totals_row, 2).value = "합계"
        ws.cell(totals_row, 2).alignment = Alignment(horizontal="center", vertical="center")
        col_letter_total = get_column_letter(col_total)
        ws.cell(totals_row, col_total).value = f"=SUM({col_letter_total}{data_start}:{col_letter_total}{last_data_row})"
        ws.cell(totals_row, col_total).number_format = "#,##0"
        ws.cell(totals_row, col_total).alignment = Alignment(horizontal="right", vertical="center")
        set_header_fill(ws, totals_row, last_col, FILL_HEADER)

        spacer_row = totals_row + 1
        set_row_border(ws, spacer_row, ws.max_column, Border())  # 무테

        # 푸터
        notice_row = last_data_row + 3
        date_row = notice_row + 1
        sign_row = notice_row + 2
        for rr in (notice_row, date_row, sign_row):
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=ws.max_column)

        ws.cell(notice_row, 1).value = "상기와 같이 내역을 확인함"
        ws.cell(notice_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        yy = year if isinstance(year, int) else datetime.now().year
        mm = month if isinstance(month, int) else datetime.now().month
        yy2, mm2 = (yy + 1, 1) if mm == 12 else (yy, mm + 1)
        ws.cell(date_row, 1).value = f"{yy2}. {mm2}."
        ws.cell(date_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        dept_str = (dept or "").strip()
        ws.cell(sign_row, 1).value = f"확인자 : {dept_str} 행정○급 ○○○ (인)"
        ws.cell(sign_row, 1).alignment = Alignment(horizontal="center", vertical="center")

        # 테두리·열너비·고정
        set_alignment(ws, range(header_row, header_row + 1), range(1, ws.max_column + 1))
        for rr in range(header_row, ws.max_row + 1):
            if rr in (spacer_row, notice_row, date_row, sign_row):
                set_row_border(ws, rr, ws.max_column, Border())
                continue
            set_row_border(ws, rr, ws.max_column, BORDER_THIN)

        auto_col_width(ws)
        for rr in range(1, ws.max_row + 1):
            ws.row_dimensions[rr].height = None
        ws.freeze_panes = ws["A6"]

    buf.seek(0)
    return buf

# ----------------------------------
# 초과근무: 분기 테이블 생성(+ 강제조정 비고·플래그)
# ----------------------------------
def _quarter_months(month: int) -> list[int]:
    q_start = ((int(month) - 1) // 3) * 3 + 1
    return [q_start, q_start + 1, q_start + 2]

def _quarter_by_qnum(q: int) -> list[int]:
    return [1,2,3] if q == 1 else [4,5,6] if q == 2 else [7,8,9] if q == 3 else [10,11,12]

def _month_col_candidates(year: int, m: int) -> list[str]:
    cands = [
        f"{m}월", f"{m:02}월",
        f"{year}-{m}", f"{year}-{m:02}",
        f"{year}.{m}", f"{year}.{m:02}",
        f"{year}/{m}", f"{year}/{m:02}",
    ]
    if m % 3 == 1:
        cands += ["분기 첫 달", "첫 달", "첫달", "분기첫달", "분기 첫 달 수당시간(h)"]
    elif m % 3 == 2:
        cands += ["분기 중간 달", "중간 달", "중간달", "분기중간달", "분기 중간 달 수당시간(h)"]
    else:
        cands += ["분기 마지막 달", "마지막 달", "마지막달", "분기마지막달", "분기 마지막 달 수당시간(h)"]
    return cands

def _pick_col(df: pd.DataFrame, names: list[str]) -> str | None:
    return find_col(df, names)

def _to_float(x) -> float:
    try:
        s = str(x).replace(",", "").strip()
        return float(s) if s != "" else 0.0
    except Exception:
        return 0.0

def _normalize_month_inputs(v1, v2, v3) -> tuple[float, float, float]:
    a, b, c = _to_float(v1), _to_float(v2), _to_float(v3)
    b_is_cum = (b >= a and c == 0 and b > (OVT_MONTH_CAP * 1.2))
    c_is_cum = (c > (a + b) and c > (OVT_MONTH_CAP * 1.5))
    if b_is_cum:
        b = max(0.0, b - a)
    if c_is_cum:
        c = max(0.0, c - (a + b))
    return a, b, c

def _fmt_g(x: float) -> str:
    return f"{x:g}"

def build_ovt_quarter_df(tmpl_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    df = tmpl_df.copy()

    c_serial = _pick_col(df, ["연번", "순번", "번호"])
    c_rank = _pick_col(df, ["직급", "직 급"])
    c_name = _pick_col(df, ["성명", "이름", "사원명"])
    if c_name is None:
        raise RuntimeError("백데이터에서 '성명' 열을 찾지 못했습니다.")

    m1, m2, m3 = _quarter_months(month)
    col_m1 = _pick_col(df, _month_col_candidates(year, m1)) or ""
    col_m2 = _pick_col(df, _month_col_candidates(year, m2)) or ""
    col_m3 = _pick_col(df, _month_col_candidates(year, m3)) or ""

    rows = []
    for _, r in df.iterrows():
        nm = str(r.get(c_name, "")).strip()
        if nm == "":
            continue
        serial = r.get(c_serial, "")
        rank = str(r.get(c_rank, "")).strip() if c_rank else ""

        raw1 = r.get(col_m1, 0)
        raw2 = r.get(col_m2, 0)
        raw3 = r.get(col_m3, 0)

        v1, v2, v3 = _normalize_month_inputs(raw1, raw2, raw3)

        a1_pre = min(v1, OVT_MONTH_CAP)
        a2_pre = min(v2, OVT_MONTH_CAP)
        a3_pre = min(v3, OVT_MONTH_CAP)

        a1 = a1_pre
        allow2 = max(0.0, OVT_QTR_CAP - a1)
        a2 = min(a2_pre, allow2)
        allow3 = max(0.0, OVT_QTR_CAP - (a1 + a2))
        a3 = min(a3_pre, allow3)

        cume1 = a1
        cume2 = a1 + a2
        cume3 = a1 + a2 + a3
        remain = max(0.0, OVT_QTR_CAP - cume3)

        adj1_m57 = (v1 > a1_pre)
        adj2_m57 = (v2 > a2_pre)
        adj3_m57 = (v3 > a3_pre)
        adj2_q90 = (a2 < a2_pre)
        adj3_q90 = (a3 < a3_pre)

        msgs = []
        if adj1_m57:
            msgs.append(f"월 57시간 초과로 시간 조정함(조정 전 : {_fmt_g(v1)} 시간)")
        if adj2_m57:
            msgs.append(f"월 57시간 초과로 시간 조정함(조정 전 : {_fmt_g(v2)} 시간)")
        if adj3_m57:
            msgs.append(f"월 57시간 초과로 시간 조정함(조정 전 : {_fmt_g(v3)} 시간)")
        if adj2_q90:
            msgs.append(f"분기 합 90시간 초과로 시간 조정함(조정 전 : {_fmt_g(a2_pre)} 시간)")
        if adj3_q90:
            msgs.append(f"분기 합 90시간 초과로 시간 조정함(조정 전 : {_fmt_g(a3_pre)} 시간)")

        note = f"잔여 가능 {int(remain)}h"
        if msgs:
            note += " / " + "; ".join(msgs)

        rows.append({
            "연번": serial,
            "직급": rank,
            "성명": nm,
            "분기 첫 달 수당시간(h)": round(a1, 2),
            "첫 달 누계(h)": round(cume1, 2),
            "분기 중간 달 수당시간(h)": round(a2, 2),
            "중간 달 누계(h)": round(cume2, 2),
            "분기 마지막 달 수당시간(h)": round(a3, 2),
            "마지막 달 누계(h)": round(cume3, 2),
            "비고": note,
            "_adj1_m57": adj1_m57,
            "_adj2_m57": adj2_m57,
            "_adj3_m57": adj3_m57,
            "_adj2_q90": adj2_q90,
            "_adj3_q90": adj3_q90,
        })

    cols = [
        "연번","직급","성명",
        "분기 첫 달 수당시간(h)","첫 달 누계(h)",
        "분기 중간 달 수당시간(h)","중간 달 누계(h)",
        "분기 마지막 달 수당시간(h)","마지막 달 누계(h)",
        "비고",
        "_adj1_m57","_adj2_m57","_adj3_m57","_adj2_q90","_adj3_q90"
    ]
    out = pd.DataFrame(rows, columns=cols)
    if "연번" in out.columns:
        out["연번"] = pd.to_numeric(out["연번"], errors="coerce").astype("Int64")
    return out

# ----------------------------------
# 초과근무 저장 서식(분기 표기 지원)
# ----------------------------------
def _rename_quarter_headers(df: pd.DataFrame, months: list[int]) -> pd.DataFrame:
    m1, m2, m3 = months
    mapping = {
        "분기 첫 달 수당시간(h)": f"{m1}월 수당시간(h)",
        "첫 달 누계(h)": f"{m1}월 누계(h)",
        "분기 중간 달 수당시간(h)": f"{m2}월 수당시간(h)",
        "중간 달 누계(h)": f"{m2}월 누계(h)",
        "분기 마지막 달 수당시간(h)": f"{m3}월 수당시간(h)",
        "마지막 달 누계(h)": f"{m3}월 누계(h)",
    }
    return df.rename(columns=mapping)

def export_ovt_view_with_format(df_full: pd.DataFrame, year: int, quarter: int, months: list[int], dept: str | None) -> BytesIO:
    visible_cols = [c for c in df_full.columns if not str(c).startswith("_")]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_full[visible_cols].to_excel(writer, sheet_name="초과근무내역", index=False, startrow=4)
        ws = writer.book["초과근무내역"]

        header_row = 5
        data_start = header_row + 1
        last_row = ws.max_row
        last_col = ws.max_column

        # 제목
        title = f"{(dept or '').strip()} 초과근무내역({year}년 {quarter}분기)"
        ws["A2"] = title
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"].font = Font(size=20, bold=True)

        set_alignment(ws, range(header_row, header_row + 1), range(1, last_col + 1))

        hdr_idx = {ws.cell(header_row, c).value: c for c in range(1, last_col + 1)}
        m1, m2, m3 = months
        cum_headers = [f"{m1}월 누계(h)", f"{m2}월 누계(h)", f"{m3}월 누계(h)"]
        cum_cols = [hdr_idx[h] for h in cum_headers if h in hdr_idx]
        col_m1 = hdr_idx.get(f"{m1}월 수당시간(h)")
        col_m2 = hdr_idx.get(f"{m2}월 수당시간(h)")
        col_m3 = hdr_idx.get(f"{m3}월 수당시간(h)")

        # 누계 강조 및 57h 표시
        for r in range(data_start, last_row + 1):
            for c in cum_cols:
                cell = ws.cell(r, c)
                cell.fill = FILL_HEADER
                cell.font = Font(bold=True)
                try:
                    val = float(str(cell.value))
                    if abs(val - 57.0) < 1e-9:
                        cell.fill = PINK
                        cell.font = Font(bold=True, color="FF0000")
                except Exception:
                    pass

        # 값이 57이면 빨강 폰트
        for r in range(data_start, last_row + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(r, c)
                try:
                    val = float(str(cell.value))
                    if abs(val - 57.0) < 1e-9:
                        cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color="FF0000",
                        )
                except Exception:
                    continue

        # 강제조정 월 빨강 Bold
        df_flags = df_full.reset_index(drop=True)
        for i in range(len(df_flags)):
            r = data_start + i
            if col_m1 and (bool(df_flags.iloc[i].get("_adj1_m57", False))):
                ws.cell(r, col_m1).font = Font(bold=True, color="FF0000")
            if col_m2 and (bool(df_flags.iloc[i].get("_adj2_m57", False)) or bool(df_flags.iloc[i].get("_adj2_q90", False))):
                ws.cell(r, col_m2).font = Font(bold=True, color="FF0000")
            if col_m3 and (bool(df_flags.iloc[i].get("_adj3_m57", False)) or bool(df_flags.iloc[i].get("_adj3_q90", False))):
                ws.cell(r, col_m3).font = Font(bold=True, color="FF0000")

        # 테두리·열너비·고정
        for rr in range(header_row, last_row + 1):
            set_row_border(ws, rr, last_col, BORDER_THIN)

        auto_col_width(ws)
        for rr in range(1, ws.max_row + 1):
            ws.row_dimensions[rr].height = None
        ws.freeze_panes = ws["A6"]

    buf.seek(0)
    return buf

# =======================================================================================
# ▲ 여기까지 ① 관내출장여비 + ② 초과근무내역 공용/로직
# ▼ 아래부터 ③ 자료 수합(머릿글 유지 수합 도구) 로직
# =======================================================================================

def _norm_cell(x) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return "" if s.lower().startswith("unnamed") else s

def _make_unique(cols):
    seen, out = {}, []
    for c in cols:
        k = (str(c) if c is not None else "").strip() or "COL"
        k = re.sub(r"\s+", " ", k)
        if k not in seen:
            seen[k] = 0
            out.append(k)
        else:
            seen[k] += 1
            out.append(f"{k}_{seen[k]}")
    return out

def _effective_header_width(head_df: pd.DataFrame) -> int:
    if head_df.empty:
        return 0
    # applymap → map (deprecation 대응)
    used = head_df.map(lambda x: bool(str(x).strip()) and str(x).lower() != "nan")
    cols_with_any = [i for i, has in enumerate(used.any(axis=0).tolist()) if has]
    return (max(cols_with_any) + 1) if cols_with_any else 0

def read_with_manual_rows(
    file_obj,
    header_first_row: int,
    header_last_row: int,
    data_start_row: int,
    data_end_row: Optional[int],
    sheet_index: int = 0,
) -> pd.DataFrame:
    raw = pd.read_excel(file_obj, sheet_name=sheet_index, header=None, dtype=str)
    if raw.empty:
        return pd.DataFrame()

    h_start = header_first_row - 1
    h_end_excl = header_last_row
    hb = raw.iloc[h_start:h_end_excl, :].copy().ffill(axis=1).ffill(axis=0).astype(str)

    ncols_eff = _effective_header_width(hb)
    if ncols_eff == 0:
        return pd.DataFrame()
    hb = hb.iloc[:, :ncols_eff]

    cols = []
    for c in range(ncols_eff):
        parts = [p.strip() for p in hb.iloc[:, c].tolist() if p and p.strip().lower() != "nan"]
        parts = [p for p in parts if not p.lower().startswith("unnamed")]
        name = re.sub(r"\s+", " ", "_".join(parts)).strip() or f"COL{c+1}"
        cols.append(name)
    cols = _make_unique(cols)

    d_start = data_start_row - 1
    d_end_excl = None if data_end_row is None else data_end_row
    body = raw.iloc[d_start:d_end_excl, :ncols_eff].copy()
    body.columns = cols
    body = body.dropna(how="all").reset_index(drop=True)
    return body

def capture_merged_header_shape_manual(
    xls_bytes: bytes,
    header_first_row: int,
    header_last_row: int,
    sheet_index: int = 0,
):
    df_all = pd.read_excel(BytesIO(xls_bytes), sheet_name=sheet_index, header=None, dtype=str)
    top = header_first_row
    bottom = header_last_row
    head = df_all.iloc[top - 1: bottom, :].copy().ffill(axis=1).ffill(axis=0).astype(str)
    ncols_eff = _effective_header_width(head)
    head = head.iloc[:, :ncols_eff].replace({"nan": "", "NaN": ""})
    head_vals = [[ _norm_cell(x) for x in head.iloc[r].tolist()] for r in range(len(head))]

    wb = load_workbook(BytesIO(xls_bytes), data_only=True)
    ws = wb.worksheets[sheet_index]
    merges_rel = []
    for rng in ws.merged_cells.ranges:
        if rng.max_row < top or rng.min_row > bottom:
            continue
        if rng.min_col > ncols_eff:
            continue
        c2 = min(rng.max_col, ncols_eff)
        r1 = rng.min_row - top + 1
        r2 = rng.max_row - top + 1
        merges_rel.append((r1, rng.min_col, r2, c2))

    return head_vals, merges_rel, ncols_eff

def write_with_merged_header_and_source(
    df: pd.DataFrame,
    head_vals,
    merges_rel,
    ncols_eff: int,
    source_col_name: str = "출처",
) -> Workbook:
    wb = Workbook()
    ws = wb.active

    hrows = len(head_vals)
    data_cols = len([c for c in df.columns if c != source_col_name])
    ncols_final = data_cols + 1  # + '출처'

    # 1) 머릿글 텍스트 채우기
    for r in range(hrows):
        row_vals = head_vals[r][:data_cols] + [""]
        for c in range(1, ncols_final + 1):
            ws.cell(r + 1, c, row_vals[c - 1] if c - 1 < len(row_vals) else "")

    # 2) 기존 병합 재적용
    limit = min(ncols_eff, data_cols)
    for (r1, c1, r2, c2) in merges_rel:
        if c1 <= limit:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=min(c2, limit))

    # 3) '출처' 머릿글 배치 및 병합
    src_col_idx = ncols_final
    ws.cell(1, src_col_idx, source_col_name)
    if hrows > 1:
        ws.merge_cells(start_row=1, start_column=src_col_idx, end_row=hrows, end_column=src_col_idx)

    # 4) 머릿글 스타일
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    font_b = Font(bold=True)
    for r in range(1, hrows + 1):
        for c in range(1, ncols_final + 1):
            cell = ws.cell(r, c)
            cell.alignment = align
            cell.font = font_b

    # 5) 본문 + 테두리 + '출처' 연파랑
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_src = PatternFill("solid", fgColor="DDEBF7")

    start_row = hrows + 1
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        # 값·서식 분리 지정 → 병합셀 value 쓰기 예외 회피
        for j, v in enumerate(row[:data_cols], start=1):
            cell = ws.cell(i, j)
            cell.value = v
            cell.border = border
        cell_src = ws.cell(i, src_col_idx)
        cell_src.value = row[data_cols]
        cell_src.fill = fill_src
        cell_src.border = border

    # 6) 열 너비 추정
    preview_end = min(start_row + max(50, len(df)), ws.max_row)
    for c in range(1, ncols_final + 1):
        maxlen = 6
        for r in range(1, preview_end + 1):
            v = ws.cell(r, c).value
            maxlen = max(maxlen, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(c)].width = min(60, maxlen + 2)

    return wb

# =======================================================================================
# 탭 UI 함수
# =======================================================================================

def tab_gwannae():
    st.title("🚗 관내출장여비 정산")
    st.markdown("📢 「인사랑」에서 결재완료된 자료를 기준으로 합니다. ")
    st.markdown("📢 새올 '차량 관리' 내역은 반영되어 있지 않습니다. ")  
    st.markdown("📢 동명이인이 존재할 경우, 에러가 발생할 수 있습니다. ")
    st.markdown("---")
    st.markdown("#### ① 업로드용 백데이터 준비（파일 열리는데 조금 걸려요）")
    st.markdown("📢 １．「인사랑」에서 관내 출장여비 엑셀을 추출해주세요．")
    if os.path.exists(MANUAL_FILE):
        with open(MANUAL_FILE, "rb") as f:
            st.download_button("📂 엑셀 추출 매뉴얼", f, file_name=MANUAL_FILE, mime="application/pdf")

    st.markdown("📢 ２． 출장자 백데이터 서식 파일입니다．")
    st.markdown("※ 연번|직급|성명|은행명|계좌번호를 입력한 후, 파일을 저장해주세요．")
    st.markdown("※ 입력된 데이터를 바탕으로, 지급조서가 생성됩니다.")
    if os.path.exists(FORM_TEMPLATE_FILE):
        with open(FORM_TEMPLATE_FILE, "rb") as f:
            st.download_button(
                "📂（서식）출장자 백데이터 파일",
                f,
                file_name=FORM_TEMPLATE_FILE,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    st.markdown("---")
    st.markdown("#### ② 파일 업로드")
    st.markdown("📢 １．「인사랑」 관내 출장여비 추출본 업로드")
    raw_up = st.file_uploader("📂 「인사랑」 관내 출장여비 추출본 업로드 (.xlsx)", type=["xlsx"], key="raw_upload")
    if raw_up is not None:
        try:
            st.session_state["RAW_DF"] = pd.read_excel(BytesIO(raw_up.getvalue()))
            st.info("✅ 「인사랑」 관내 출장여비 원본 업로드 완료")
        except Exception as e:
            st.error(f"🚫 「인사랑」 관내 출장여비 파일 읽기 오류: {e}")

    st.markdown("📢 ２．출장자 백데이터 업로드")
    tmpl_up = st.file_uploader("📂 출장자 백데이터 업로드 (.xlsx)", type=["xlsx"], key="tmpl_upload")
    if tmpl_up is not None:
        try:
            st.session_state["TMPL_DF"] = read_template_dataframe(BytesIO(tmpl_up.getvalue()))
            st.info("✅ 출장자 백데이터 업로드 완료")
        except Exception as e:
            st.error(f"🚫 출장자 백데이터 읽기 오류: {e}")

    st.markdown("---")
    st.markdown("#### ③ 데이터 가공 · 요약")
    st.markdown("📢 부서명을 입력하고, 필요시 날짜 포함/제외 규칙을 추가하세요.")
    btn = st.button("⌛ 가공 실행(백데이터→가공→요약)", type="primary", disabled=(raw_up is None))
    if btn:
        try:
            with st.spinner("처리 중..."):
                wb = prepare_backend_sheet_xlsx(BytesIO(raw_up.getvalue()))
                wb, proc_df = create_gagong_and_summary(wb)
                st.session_state["PROC_DF"] = proc_df
                st.session_state["OUT_BYTES"] = save_wb_to_bytes(wb)
            st.success("가공이 완료되었습니다.")
            st.download_button(
                "💾 요약 결과 엑셀 다운로드",
                data=st.session_state["OUT_BYTES"],
                file_name=f"관내출장_가공요약_{kst_timestamp()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        except Exception as e:
            st.error(f"오류: {e}")

    if "PROC_DF" in st.session_state:
        st.markdown("##### 요약 편집")
        if "ADJUST_RULES" not in st.session_state:
            st.session_state["ADJUST_RULES"] = {}

        df = st.session_state["PROC_DF"].copy()
        df["도착일자_dt"] = df["도착일자"].apply(to_datetime_flex)
        df["지급단가"] = pd.to_numeric(df["지급단가"], errors="coerce").fillna(0).astype(int)
        df = df[(df["출장자"].astype(str).str.strip() != "") & (~df["도착일자_dt"].isna())]
        if df.empty:
            st.info("표시할 데이터가 없습니다.")
            return

        years_in_data = sorted(df["도착일자_dt"].dt.year.dropna().unique().tolist())
        base_years = years_in_data if years_in_data else [datetime.now().year]
        min_y, max_y = min(base_years), max(base_years)
        year_options = sorted(set(base_years + [min_y - 1, max_y + 1]))
        default_year = datetime.now().year if datetime.now().year in year_options else max(base_years)

        dept_name = st.text_input("부서명", value=st.session_state.get("DEPT_NAME", ""), key="dept_name")
        st.session_state["DEPT_NAME"] = dept_name

        cY, cM = st.columns([1, 1])
        with cY:
            sel_year = st.selectbox("출장연도", options=year_options,
                                    index=year_options.index(default_year), key="yr_sel")
        months_in_year = sorted(df[df["도착일자_dt"].dt.year == sel_year]["도착일자_dt"].dt.month.dropna().unique().tolist())
        month_options = list(range(1, 12 + 1))
        default_month = (months_in_year[-1] if months_in_year else datetime.now().month)
        with cM:
            sel_month = st.selectbox("출장월", options=month_options,
                                     index=month_options.index(default_month), key="mo_sel")

        df_ym = df[(df["도착일자_dt"].dt.year == sel_year) & (df["도착일자_dt"].dt.month == sel_month)]
        if df_ym.empty:
            st.info("선택한 연·월 데이터가 없습니다.")
            return

        base_dates: dict[tuple[str, int], list] = {}
        for (nm, pay), grp in df_ym.groupby(["출장자", "지급단가"]):
            base_dates[(str(nm), int(pay))] = sorted({d.date() for d in grp["도착일자_dt"]})
        names_all = sorted({nm for nm, _ in base_dates.keys()})

        c1, c2, c3, c4 = st.columns([1.6, 1.2, 1.0, 3.0])
        with c1:
            sel_name = st.selectbox("출장자", names_all, key="name_sel")
        with c2:
            pays_of_name = sorted({pay for (nm, pay) in base_dates.keys() if nm == sel_name}, reverse=True)
            sel_pay = st.selectbox("지급단가", pays_of_name, key="pay_sel")
        with c3:
            mode = st.radio("모드", ["제외", "포함"], horizontal=True, key="mode_sel")
        with c4:
            pool_dates = [d.strftime("%Y-%m-%d") for d in base_dates.get((sel_name, int(sel_pay)), [])]
            chosen = st.multiselect("날짜 선택", options=pool_dates, default=[], key="dates_sel")

        b1, b2 = st.columns([1, 1])
        with b1:
            add_clicked = st.button("➕ 추가", width="stretch")
        with b2:
            reset_clicked = st.button("🔄 초기화", width="stretch")

        if add_clicked:
            if chosen:
                key = (sel_name, int(sel_pay))
                st.session_state["ADJUST_RULES"][key] = {
                    "mode": mode,
                    "dates": {datetime.strptime(s, "%Y-%m-%d").date() for s in chosen},
                }
                st.success(f"규칙 저장: {sel_name} / {sel_pay:,}원 / {mode} / {len(chosen)}개")
            else:
                st.warning("날짜를 선택하세요.")
        if reset_clicked:
            st.session_state["ADJUST_RULES"] = {}
            st.info("누적 규칙을 초기화했습니다.")

        included_map: dict[tuple[str, int], list] = {}
        adj = st.session_state["ADJUST_RULES"]
        for key, days in base_dates.items():
            if key in adj:
                a = adj[key]
                labels_all = set(days)
                chosen_set = set(a["dates"])
                included_map[key] = sorted(list(labels_all - chosen_set)) if a["mode"] == "제외" \
                    else sorted(list(labels_all & chosen_set))
            else:
                included_map[key] = sorted(days)

        rows = []
        for (nm, pay) in sorted(base_dates.keys(), key=lambda x: (x[0], -x[1])):
            dd = included_map.get((nm, pay), [])
            rows.append({
                "성명": nm,
                "지급단가": int(pay),
                "출장일수": len(dd),
                "여비합계": int(pay) * len(dd),
                "출장현황": ", ".join([str(x.day) for x in dd]),
            })
        summary_all = pd.DataFrame(rows, columns=["성명", "지급단가", "출장일수", "여비합계", "출장현황"])

        st.dataframe(summary_all, width="stretch")
        cA, cB, cC = st.columns(3)
        with cA:
            st.metric("총 인원", f"{summary_all['성명'].nunique()}")
        with cB:
            st.metric("총 출장일수", f"{int(summary_all['출장일수'].sum())}")
        with cC:
            st.metric("총 소계", f"{int(summary_all['여비합계'].sum()):,} 원")

        st.session_state["SUMMARY_RESULT_DF"] = summary_all
        st.session_state["SUMMARY_YEAR"] = sel_year
        st.session_state["SUMMARY_MONTH"] = sel_month

        st.markdown("---")
        st.markdown("#### ④ 지급 조서 다운로드")

        disabled = ("TMPL_DF" not in st.session_state or st.session_state.get("TMPL_DF", pd.DataFrame()).empty)
        if disabled:
            st.info("혼합 내보내기를 하려면 (서식) 출장자 백데이터를 업로드하세요.")
        else:
            try:
                mixed_df = build_mixed_df(summary_all, st.session_state["TMPL_DF"])
                xbytes = export_mixed_to_excel(
                    mixed_df,
                    st.session_state.get("SUMMARY_YEAR"),
                    st.session_state.get("SUMMARY_MONTH"),
                    st.session_state.get("DEPT_NAME", ""),
                )

                dept = (st.session_state.get("DEPT_NAME") or "").strip() or "부서미지정"
                year = st.session_state.get("SUMMARY_YEAR")
                month = st.session_state.get("SUMMARY_MONTH")
                fname = f"{dept} 관내출장여비_지급조서({year}년 {month}월).xlsx"

                st.download_button(
                    "💾 지급 조서 다운로드",
                    data=xbytes,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                )
                st.dataframe(mixed_df, width="stretch", height=360)

            except Exception as e:
                st.error(f"지급 조서 생성 오류: {e}")

def tab_overtime():
    st.title("⏱️ 초과근무내역")
    st.markdown("---")

    st.markdown("#### ① 업로드용 백데이터 준비（파일 열리는데 조금 걸려요）")
    st.markdown("📢 초과근무자 백데이터 서식 파일입니다．")
    st.markdown("※ 연번|직급|성명|「인사랑」에서 조회되는 초과수당시간을 입력하세요.")
    st.markdown("※ 데이터 입력 시, 머릿글(헤더)은 수정하시면 안됩니다. 🚫 ")
    if os.path.exists(FORM_TEMPLATE_FILE_OVT):
        with open(FORM_TEMPLATE_FILE_OVT, "rb") as f:
            st.download_button(
                "📂（서식）초과근무자 백데이터",
                f,
                file_name=FORM_TEMPLATE_FILE_OVT,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    st.markdown("---")
    st.markdown("#### ② 파일 업로드")
    st.markdown("📢 작성 완료한 ‘초과근무자 백데이터’ 파일을 업로드해주세요．")
    st.markdown("※ 백데이터 파일을 토대로, 엑셀 파일이 생성됩니다．")

    tmpl_up = st.file_uploader("📂 초과근무자 백데이터 업로드 (.xlsx)", type=["xlsx"], key="ovt_tmpl_upload")
    if tmpl_up is not None:
        try:
            st.session_state["OVT_TMPL_DF"] = read_template_dataframe(BytesIO(tmpl_up.getvalue()))
            st.info("✅ 초과근무자 백데이터 업로드 완료")
        except Exception as e:
            st.error(f"🚫 초과근무자 백데이터 읽기 오류: {e}")

    st.markdown("---")
    st.markdown("#### ③ 데이터 가공 · 요약")
    st.markdown("📢 부서명과 기준 연·분기를 입력하세요.")
    dept_name = st.text_input("부서명", value=st.session_state.get("OVT_DEPT_NAME", ""), key="ovt_dept_name")
    st.session_state["OVT_DEPT_NAME"] = dept_name

    cY, cQ = st.columns([1, 1])
    with cY:
        sel_year = st.number_input("초과근무 연도", min_value=2000, max_value=2100,
                                   value=int(st.session_state.get("OVT_YEAR", datetime.now().year)),
                                   step=1, key="ovt_year_in")
    with cQ:
        q_labels = {1: "1분기(1~3월)", 2: "2분기(4~6월)", 3: "3분기(7~9월)", 4: "4분기(10~12월)"}
        q_options = [1, 2, 3, 4]
        default_q = int(st.session_state.get("OVT_QTR", ((datetime.now().month - 1)//3)+1))
        sel_quarter = st.selectbox("초과근무 분기", options=q_options,
                                   index=q_options.index(default_q),
                                   format_func=lambda x: q_labels[x],
                                   key="ovt_quarter_in")

    btn = st.button("⌛ 가공 실행(백데이터→분기테이블)", type="primary",
                    disabled=("OVT_TMPL_DF" not in st.session_state or st.session_state["OVT_TMPL_DF"].empty))

    if btn:
        try:
            with st.spinner("처리 중..."):
                months = _quarter_by_qnum(int(sel_quarter))         # [m1, m2, m3]
                ref_month = months[0]                               # 내부 계산용 기준월
                df_quarter = build_ovt_quarter_df(st.session_state["OVT_TMPL_DF"], int(sel_year), int(ref_month))
                df_quarter_named = _rename_quarter_headers(df_quarter, months)
                view_df = df_quarter_named.drop(columns=[c for c in df_quarter_named.columns if str(c).startswith("_")])

                st.dataframe(view_df, width="stretch")
                st.session_state["OVT_Q_DF"] = df_quarter_named
                st.session_state["OVT_VIEW_DF"] = view_df
                st.session_state["OVT_YEAR"] = int(sel_year)
                st.session_state["OVT_QTR"] = int(sel_quarter)
                st.session_state["OVT_Q_MONTHS"] = months
        except Exception as e:
            st.error(f"오류: {e}")

    st.markdown("---")
    st.markdown("#### ④ 초과근무내역 엑셀 저장")
    st.markdown("📢 상한시간을 초과한 경우, 빨간색으로 표시하였습니다.")
    st.markdown("📢 수기 또는 재난 비상근무 시간은 제외된 자료이므로, 참고용으로 활용하세요.")
    if "OVT_Q_DF" not in st.session_state:
        st.info("③ 가공을 먼저 실행하세요.")
    else:
        try:
            dept = (st.session_state.get("OVT_DEPT_NAME") or "").strip() or "부서미지정"
            year = int(st.session_state.get("OVT_YEAR"))
            quarter = int(st.session_state.get("OVT_QTR"))
            months = st.session_state.get("OVT_Q_MONTHS", _quarter_by_qnum(quarter))

            fname = f"{dept} 초과근무내역({year}년 {quarter}분기).xlsx"
            xbytes = export_ovt_view_with_format(st.session_state["OVT_Q_DF"], year, quarter, months, dept)
            st.download_button(
                "💾 초과근무내역 엑셀 저장",
                data=xbytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        except Exception as e:
            st.error(f"엑셀 생성 오류: {e}")

def tab_collect():
    st.title("📊 자료 수합(엑셀 파일만 가능)")
    st.markdown("📢 각 수합 파일의 머릿글은 모두 동일해야 합니다. ")
    st.markdown("📢 수합 데이터 범위 내 행과 행 사이에 빈 행이 존재하면 안됩니다. ")
    st.markdown("📢 수합 데이터 범위 내 병합된 셀들이 존재하면 안됩니다. ")  
    st.markdown("---")

    # ① 업로드
    st.markdown("### ① 수합 대상 파일 업로드")
    files = st.file_uploader("※ 엑셀 파일 복수 선택 가능", type=["xls", "xlsx"], accept_multiple_files=True)

    # ② 머릿글 범위(세로)
    st.markdown("### ② 머릿글(헤더) 범위 설정")
    st.markdown("📢 수합 파일 내 머릿글의 범위를 먼저 확인해주세요")
    header_first = st.number_input("머릿글 첫 행", min_value=1, value=1, step=1)   # 기본 1
    header_last  = st.number_input("머릿글 마지막 행", min_value=1, value=1, step=1)  # 기본 1
    st.caption("예) 머릿글이 1~2행이면 첫 행=1, 마지막 행=2")
    st.caption("예) 머릿글이   1행이면 첫 행=1, 마지막 행=1")

    # ③ 데이터 범위(세로)
    st.markdown("### ③ 수합 데이터 범위 설정")
    st.markdown("📢 수합 파일 내 수합 데이터 범위를 먼저 확인해주세요")
    data_start = st.number_input("데이터 시작 행", min_value=1, value=2, step=1)
    to_end = st.checkbox("데이터 마지막 행 = 해당 시트 맨 아래 끝까지", value=True)
    data_end = None
    if not to_end:
        data_end = st.number_input("데이터 마지막 행", min_value=1, value=max(2, data_start), step=1)

    st.caption("데이터 마지막 행을 직접 지정하고 싶으시면, '체크 해제'하세요")
    st.caption("데이터 범위 지정은 업로드된 파일들에 공통으로 적용됩니다.")

    # 실행
    if files:
        if header_last < header_first:
            st.error("머릿글 마지막 행은 머릿글 첫 행보다 크거나 같아야 합니다.")
            st.stop()
        if data_start <= header_last:
            st.error("데이터 시작 행은 머릿글 마지막 행보다 커야 합니다.")
            st.stop()
        if data_end is not None and data_end < data_start:
            st.error("데이터 마지막 행은 데이터 시작 행보다 크거나 같아야 합니다.")
            st.stop()

        blobs = [f.read() for f in files]
        names = [f.name for f in files]

        head_vals, merges_rel, ncols_eff = capture_merged_header_shape_manual(
            blobs[0], header_first_row=header_first, header_last_row=header_last
        )

        dfs = []
        for name, b in zip(names, blobs):
            df = read_with_manual_rows(BytesIO(b), header_first, header_last, data_start, data_end)
            if not df.empty:
                df["출처"] = name
                cols = [c for c in df.columns if c != "출처"] + ["출처"]
                df = df[cols]
                dfs.append(df)

        if not dfs:
            st.error("수합 가능한 데이터가 없습니다.")
        else:
            merged = pd.concat(dfs, ignore_index=True)

            wb = write_with_merged_header_and_source(
                merged, head_vals, merges_rel, ncols_eff, source_col_name="출처"
            )
            buf = BytesIO(); wb.save(buf); buf.seek(0)

            st.markdown("### ④ 수합 완료본 다운로드")
            ts = _kst_now().strftime("%y%m%d_%H%M")  # yymmdd_hhmm
            out_name = f"수합 완료본_{ts}.xlsx"
            st.download_button(
                "💾 수합 완료본 다운로드",
                data=buf.getvalue(),
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

            st.success(f"수합 완료: {len(dfs)}개 파일, {len(merged):,}행")
            st.dataframe(merged.head(50), width="stretch")
    else:
        st.info("엑셀 파일을 업로드하세요.")

# ----------------------------------
# 메인
# ----------------------------------
def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)
    tabs = st.tabs(["관내출장여비", "초과근무내역", "자료 수합"])
    with tabs[0]:
        tab_gwannae()
    with tabs[1]:
        tab_overtime()
    with tabs[2]:
        tab_collect()

if __name__ == "__main__":
    main()



